"""
Orange County Housing Justice Dashboard
Paxton Wang — 2026
Run: python3 -m streamlit run oc_housing_dashboard.py
"""

import os
import json
import random
import urllib.request

import streamlit as st
import pandas as pd
import numpy as np
import folium
from streamlit_folium import st_folium
import plotly.express as px
import plotly.graph_objects as go

# ── PAGE CONFIG ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="OC Housing Justice",
    page_icon="🏠",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── GLOBAL CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .stApp, [data-testid="stAppViewContainer"], .main, .block-container {
        background-color: #F7FAFD !important;
    }
    html, body, p, span, div, label, li, td, th, a,
    .stMarkdown, .stMarkdown p, .stMarkdown span,
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] span {
        color: #1A2B3C !important;
    }
    h1, h2, h3, h4, h5 { color: #1A5276 !important; }
    [data-testid="stSidebar"] { background-color: #EAF4FB !important; }
    [data-testid="stSidebar"] * { color: #1A2B3C !important; }
    .stRadio label, .stCheckbox label, .stSelectbox label,
    .stMultiselect label, .stSlider label, .stNumberInput label { color: #1A2B3C !important; }
    .metric-card {
        background: white !important; border-radius: 10px; padding: 20px;
        box-shadow: 0 2px 10px rgba(26,82,118,0.10);
        border-left: 5px solid #1A5276; margin-bottom: 16px;
    }
    .metric-value { font-size: 2.2rem; font-weight: 700; color: #1A5276 !important; }
    .metric-label { font-size: 0.9rem; color: #5D8AA8 !important; margin-top: 4px; }
    .insight-box {
        background: #EAF4FB !important; border-left: 5px solid #2E86C1;
        border-radius: 8px; padding: 16px; margin: 12px 0;
    }
    .insight-box, .insight-box * { color: #1A2B3C !important; }
    .alert-box {
        background: #FEF0E6 !important; border-left: 5px solid #D95A00;
        border-radius: 8px; padding: 16px; margin: 10px 0;
    }
    .alert-box, .alert-box * { color: #1A2B3C !important; }
    .alert-box h4 { color: #D95A00 !important; }
</style>
""", unsafe_allow_html=True)

# ── DATA PATH ──────────────────────────────────────────────────────────────────
DATA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "HousingCentersOC.xlsx")

# ── LOAD DATA ──────────────────────────────────────────────────────────────────
@st.cache_data
def load_data(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[2]]
    data = [row for row in ws.iter_rows(min_row=3, values_only=True)]
    df = pd.DataFrame(data, columns=headers)

    numeric_cols = [
        'Total Beds', 'Rent Burden %', 'Median HH Income Percentile',
        'Poverty Line Percentile', 'CalEnviroScreen 4.0 Percentile',
        'Pit Count', 'Utilization Rate', 'Year-Round Beds',
        'Beds HH w/ Children', 'Beds HH w/o Children'
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    df = df.dropna(subset=['Project Type', 'City'], how='all').reset_index(drop=True)

    key_vars = ['Rent Burden %', 'Median HH Income Percentile', 'Poverty Line Percentile']
    enriched = df[df['Census Tract'].notna()].copy()
    enriched_clean = enriched.dropna(subset=key_vars).copy()
    for col in key_vars + ['Total Beds']:
        enriched_clean[col] = pd.to_numeric(enriched_clean[col], errors='coerce')

    enriched_clean['Income Need'] = 1 - enriched_clean['Median HH Income Percentile']
    enriched_clean['Need Score'] = (
        enriched_clean['Rent Burden %'].astype(float) * 0.40 +
        enriched_clean['Income Need'].astype(float) * 0.25 +
        enriched_clean['Poverty Line Percentile'].astype(float) * 0.35
    ).round(3)

    tract_agg = enriched_clean.groupby(['Census Tract', 'City']).agg(
        Total_Beds=('Total Beds', 'sum'),
        Num_Programs=('Project Type', 'count'),
        Rent_Burden=('Rent Burden %', 'mean'),
        Income_Percentile=('Median HH Income Percentile', 'mean'),
        Poverty_Percentile=('Poverty Line Percentile', 'mean'),
        Enviro_Percentile=('CalEnviroScreen 4.0 Percentile', 'mean'),
        Need_Score=('Need Score', 'mean'),
        Project_Types=('Project Type', lambda x: ', '.join(x.dropna().unique()))
    ).reset_index().round(3)

    med_beds = tract_agg['Total_Beds'].median()
    med_need = tract_agg['Need_Score'].median()
    tract_agg['Underserved'] = (
        (tract_agg['Need_Score'] > med_need) & (tract_agg['Total_Beds'] < med_beds)
    )

    city_summary = enriched_clean.groupby('City').agg(
        Total_Beds=('Total Beds', 'sum'),
        Avg_Need_Score=('Need Score', 'mean'),
        Num_Programs=('Project Type', 'count'),
        Avg_Rent_Burden=('Rent Burden %', 'mean'),
        Avg_Poverty=('Poverty Line Percentile', 'mean')
    ).reset_index().round(3)

    return df, enriched_clean, tract_agg, city_summary


@st.cache_data
def load_tract_geojson():
    geojson_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "oc_tracts.geojson")
    if os.path.exists(geojson_path):
        with open(geojson_path, 'r') as f:
            return json.load(f)
    urls = [
        (
            "https://tigerweb.geo.census.gov/arcgis/rest/services/TIGERweb/"
            "tigerWMS_Census2020/MapServer/8/query"
            "?where=STATE%3D06+AND+COUNTY%3D059"
            "&outFields=TRACT%2CGEOID%2CNAME%2CNAMELSAD"
            "&returnGeometry=true&f=geojson&resultRecordCount=600"
        ),
    ]
    for url in urls:
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=25) as r:
                data = json.loads(r.read())
            if data.get("features"):
                with open(geojson_path, 'w') as f:
                    json.dump(data, f)
                return data
        except Exception:
            continue
    return None


def load_ces50():
    """Load CalEnviroScreen 5.0 percentiles for all OC tracts."""
    candidates = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "ces50_oc.json"),
        os.path.join(os.getcwd(), "ces50_oc.json"),
        os.path.expanduser("~/Desktop/ces50_oc.json"),
        "ces50_oc.json",
    ]
    for path in candidates:
        if os.path.exists(path):
            with open(path, "r") as f:
                return json.load(f)
    return {}


df, enriched, tract_agg, city_summary = load_data(DATA_PATH)
tract_geojson = load_tract_geojson()
ces50 = load_ces50()

# ── HELPERS ────────────────────────────────────────────────────────────────────
AXIS_STYLE = dict(
    color='#1A2B3C',
    tickfont=dict(color='#1A2B3C', size=12),
    title_font=dict(color='#1A2B3C'),
    gridcolor='#E8EEF4',
    linecolor='#CCDDEE'
)


def style_fig(fig):
    fig.update_layout(
        paper_bgcolor='white',
        plot_bgcolor='white',
        font=dict(color='#1A2B3C', size=13, family='sans-serif'),
    )
    fig.update_xaxes(**AXIS_STYLE)
    fig.update_yaxes(**AXIS_STYLE)
    fig.update_coloraxes(colorbar=dict(
        tickfont=dict(color='#1A2B3C'),
        title_font=dict(color='#1A2B3C'),
    ))
    return fig


def need_color(score):
    if score >= 0.80: return '#D95A00'
    if score >= 0.70: return '#E8870A'
    if score >= 0.60: return '#F0A500'
    return '#2E86C1'


def clean_tract(t):
    s = str(t).strip()
    if '.' in s:
        s = s.split('.')[0]
    return s.zfill(6)


def enviro_color(val):
    if val is None or (isinstance(val, float) and val != val):
        return '#AAAAAA'
    t = float(val)
    stops = [
        (0.00, (0,   80,  0)),
        (0.10, (32, 128,  0)),
        (0.20, (64, 160,  0)),
        (0.30, (128,192,  0)),
        (0.40, (208,208,  0)),
        (0.50, (240,208,  0)),
        (0.60, (240,160,  0)),
        (0.70, (240,112,  0)),
        (0.80, (224, 64,  0)),
        (0.90, (192,  0,  0)),
        (1.00, (160,  0,  0)),
    ]
    for i in range(len(stops) - 1):
        t0, c0 = stops[i]
        t1, c1 = stops[i + 1]
        if t0 <= t <= t1:
            frac = (t - t0) / (t1 - t0) if t1 > t0 else 0
            return '#{:02X}{:02X}{:02X}'.format(
                int(c0[0] + (c1[0] - c0[0]) * frac),
                int(c0[1] + (c1[1] - c0[1]) * frac),
                int(c0[2] + (c1[2] - c0[2]) * frac),
            )
    return '#AAAAAA'


def lerp_color(val, c0, c1):
    if val is None or (isinstance(val, float) and val != val):
        return '#AAAAAA'
    t = float(val)
    return '#{:02X}{:02X}{:02X}'.format(
        int(c0[0] + (c1[0] - c0[0]) * t),
        int(c0[1] + (c1[1] - c0[1]) * t),
        int(c0[2] + (c1[2] - c0[2]) * t),
    )


# ── TRACT COORDINATES (Census TIGERweb 2020, 482 OC tracts) ───────────────────
TRACT_COORDS = {
    '001101': (33.9429662, -117.9736062), '001102': (33.9381467, -117.9723923),
    '001201': (33.9339694, -117.9555604), '001202': (33.9347338, -117.9396017),
    '001301': (33.920151,  -117.9626323), '001303': (33.9193422, -117.9553075),
    '001401': (33.9415329, -117.9482808), '001402': (33.9355599, -117.9267923),
    '001403': (33.9429298, -117.919917),  '001404': (33.9258647, -117.9361063),
    '001501': (33.9282466, -117.9199855), '001503': (33.917555,  -117.9114788),
    '001504': (33.9158797, -117.8938686), '001505': (33.9087038, -117.8984354),
    '001506': (33.9345638, -117.8845424), '001507': (33.9249188, -117.8893765),
    '001602': (33.8903105, -117.8956295), '001604': (33.9142976, -117.9246065),
    '001704': (33.8898869, -117.9482615), '001705': (33.9156337, -117.9394188),
    '001706': (33.9036335, -117.9347908), '001708': (33.9100483, -117.9534349),
    '001710': (33.9014105, -117.9641014), '001801': (33.8620002, -117.9828428),
    '001802': (33.8682303, -117.972226),  '001901': (33.8628285, -117.9658925),
    '011000': (33.8675559, -117.9465935), '011101': (33.8568109, -117.9458335),
    '011200': (33.8775171, -117.9353698), '011403': (33.8723103, -117.9026631),
    '011502': (33.8708185, -117.8847443), '011503': (33.8759477, -117.8939652),
    '011601': (33.8654371, -117.926416),  '011602': (33.8634067, -117.9220316),
    '011707': (33.8984401, -117.8827253), '011709': (33.9001161, -117.8593918),
    '011710': (33.8931934, -117.8676897), '011711': (33.8876233, -117.8701081),
    '011712': (33.8755983, -117.8659429), '011715': (33.8865549, -117.8484028),
    '011717': (33.9009709, -117.8520628), '011718': (33.8938867, -117.8514203),
    '021802': (33.8912918, -117.8070053), '021807': (33.8691264, -117.7667933),
    '021810': (33.8842463, -117.828946),  '021812': (33.867715,  -117.8116429),
    '021813': (33.8591647, -117.8165774), '021814': (33.9218513, -117.8695161),
    '021816': (33.8713134, -117.7944769), '021817': (33.8749236, -117.7854332),
    '021821': (33.8659379, -117.8311042), '021822': (33.9115917, -117.7802177),
    '021823': (33.896973,  -117.7814562), '021824': (33.9012368, -117.7681073),
    '021825': (33.8850132, -117.7503714), '021826': (33.8758129, -117.7290323),
    '021827': (33.8797325, -117.7269325), '021828': (33.8857369, -117.7145436),
    '021830': (33.8813512, -117.768185),  '021831': (33.916221,  -117.8336153),
    '021832': (33.9347931, -117.8623627), '021903': (33.847166,  -117.828601),
    '021905': (33.8653809, -117.7635213), '021912': (33.8033892, -117.7635349),
    '021913': (33.7893669, -117.8055462), '021914': (33.77808,   -117.806058),
    '021915': (33.8390215, -117.8251231), '021916': (33.8405137, -117.7872697),
    '021917': (33.8084735, -117.7765861), '021918': (33.792008,  -117.7905564),
    '021920': (33.8381855, -117.7490113), '021921': (33.8541937, -117.7492002),
    '021922': (33.8550094, -117.7407222), '021923': (33.8306216, -117.7518979),
    '021924': (33.8618738, -117.7229019),
    '032002': (33.6138461, -117.6720523), '032003': (33.6037108, -117.6597465),
    '032011': (33.7165885, -117.5914249), '032013': (33.5649979, -117.6675715),
    '032014': (33.622033,  -117.6926346), '032015': (33.6040618, -117.6838397),
    '032020': (33.6030923, -117.6462868), '032022': (33.5353355, -117.670529),
    '032027': (33.6298107, -117.6846851), '032028': (33.6213784, -117.6781362),
    '032029': (33.642218,  -117.6657857), '032030': (33.6487138, -117.6552477),
    '032031': (33.6436834, -117.6503733), '032032': (33.6320278, -117.6585663),
    '032033': (33.6484517, -117.6419026), '032034': (33.6461095, -117.6322208),
    '032035': (33.6330078, -117.6448716), '032039': (33.5732231, -117.6502696),
    '032040': (33.5770935, -117.6547617), '032042': (33.6459537, -117.572819),
    '032043': (33.5862353, -117.5686313), '032044': (33.6057579, -117.5768416),
    '032045': (33.5859851, -117.5937305), '032046': (33.5754299, -117.5791707),
    '032047': (33.6578404, -117.6350957), '032048': (33.6542953, -117.6395692),
    '032049': (33.6603442, -117.614871),  '032050': (33.6468675, -117.6016471),
    '032051': (33.6518941, -117.5838495), '032053': (33.5912899, -117.6323947),
    '032054': (33.640619,  -117.5886183), '032055': (33.6450858, -117.5856654),
    '032057': (33.5613888, -117.647131),  '032059': (33.5579172, -117.6412882),
    '032061': (33.5182818, -117.6579773), '032062': (33.4735859, -117.6046879),
    '032063': (33.4660371, -117.5794885), '032065': (33.6238873, -117.6052805),
    '032066': (33.5352892, -117.5972598),
    '042106': (33.437821,  -117.6353091), '042107': (33.4298606, -117.6226617),
    '042108': (33.4273086, -117.6163043), '042109': (33.4428718, -117.5986285),
    '042111': (33.479822,  -117.6169906), '042112': (33.4536983, -117.6247318),
    '042114': (33.4124915, -117.5942747), '042115': (33.3977143, -117.6006656),
    '042116': (33.4139587, -117.6046144), '042201': (33.4628582, -117.670744),
    '042206': (33.4512478, -117.6495561), '042305': (33.5029456, -117.7399067),
    '042312': (33.5066758, -117.6672036), '042317': (33.5040533, -117.7155868),
    '042319': (33.5418207, -117.7107447), '042320': (33.5771822, -117.708596),
    '042323': (33.4848853, -117.7163209), '042324': (33.4879016, -117.721039),
    '042325': (33.5278504, -117.7323556), '042326': (33.5391635, -117.7189072),
    '042327': (33.5914751, -117.6914181), '042329': (33.526655,  -117.7039693),
    '042330': (33.5261729, -117.6919402), '042331': (33.5190266, -117.6896548),
    '042332': (33.5045656, -117.7015571), '042333': (33.5664995, -117.6844837),
    '042335': (33.5621941, -117.6990587), '042337': (33.5125895, -117.6857019),
    '042339': (33.4745668, -117.6979879), '042341': (33.4908118, -117.6690498),
    '052408': (33.6451436, -117.70455),   '052410': (33.6333487, -117.7052789),
    '052415': (33.6374069, -117.6989554), '052416': (33.6329082, -117.691144),
    '052417': (33.7087908, -117.7559728), '052421': (33.7203564, -117.764267),
    '052422': (33.6786702, -117.6670441), '052423': (33.6581054, -117.6683),
    '052424': (33.6460253, -117.6723393), '052425': (33.6591282, -117.6825894),
    '052427': (33.6843238, -117.6482724), '052428': (33.6801011, -117.6360687),
    '052429': (33.7046013, -117.7370286), '052430': (33.696581,  -117.657368),
    '052431': (33.7026155, -117.7503042), '052433': (33.6892167, -117.7639414),
    '052434': (33.740563,  -117.7707091), '052435': (33.7254828, -117.7734343),
    '052436': (33.7239383, -117.7489524), '052437': (33.7351576, -117.7553263),
    '052438': (33.682358,  -117.7084222), '052439': (33.6822596, -117.7386387),
    '052502': (33.7146018, -117.8101091), '052505': (33.6928736, -117.7738735),
    '052506': (33.6879916, -117.7831853), '052511': (33.6825257, -117.7996196),
    '052513': (33.6765051, -117.7940675), '052514': (33.6716949, -117.7927631),
    '052518': (33.6508001, -117.7444948), '052519': (33.6673361, -117.8111762),
    '052520': (33.6768063, -117.8042215), '052521': (33.6839459, -117.8157827),
    '052522': (33.6822972, -117.8231376), '052523': (33.6753876, -117.8296673),
    '052524': (33.727085,  -117.7990388), '052526': (33.7022092, -117.784018),
    '052527': (33.6999175, -117.7971039), '052528': (33.6904484, -117.7898794),
    '052530': (33.6874139, -117.8088534), '052531': (33.6655734, -117.7865333),
    '052533': (33.7090063, -117.7740975), '052535': (33.7214669, -117.7780293),
    '062610': (33.6660688, -117.8516041), '062614': (33.6366742, -117.8401467),
    '062619': (33.5458658, -117.764034),  '062620': (33.5245325, -117.7687725),
    '062622': (33.6246547, -117.7243279), '062625': (33.5908167, -117.7090144),
    '062626': (33.650382,  -117.8380958), '062629': (33.6468099, -117.8211337),
    '062630': (33.6370203, -117.7986393), '062631': (33.6408111, -117.8179363),
    '062632': (33.5492705, -117.7499543), '062633': (33.558816,  -117.7381176),
    '062634': (33.5929201, -117.717666),  '062635': (33.5991037, -117.7423379),
    '062636': (33.5840081, -117.7144774), '062637': (33.5771513, -117.7266318),
    '062638': (33.5766786, -117.7359323), '062639': (33.5663099, -117.7321437),
    '062641': (33.6040101, -117.7336606), '062642': (33.6049352, -117.8655656),
    '062643': (33.5917535, -117.8128852), '062645': (33.6200531, -117.8359983),
    '062649': (33.5967106, -117.7558176), '062650': (33.6617274, -117.8142617),
    '062651': (33.6593178, -117.8043563), '062652': (33.6262805, -117.7412855),
    '062653': (33.6428868, -117.756721),  '062654': (33.6450713, -117.7675029),
    '062655': (33.6299187, -117.8358616), '062656': (33.5473661, -117.798695),
    '062657': (33.6331971, -117.8610092), '062658': (33.6169148, -117.8600016),
    '062701': (33.6018474, -117.8756986), '062702': (33.5900444, -117.870786),
    '063005': (33.6130439, -117.8955202), '063007': (33.6236938, -117.8688714),
    '063008': (33.6213206, -117.8784353), '063009': (33.6524313, -117.8789087),
    '063010': (33.6253897, -117.9077538), '063101': (33.6646109, -117.8854456),
    '063201': (33.6480393, -117.9063375), '063202': (33.6442198, -117.9079689),
    '063301': (33.6370234, -117.9142547), '063302': (33.62929,   -117.9201115),
    '063400': (33.6167825, -117.9153852), '063500': (33.6159935, -117.9393261),
    '063603': (33.6258225, -117.9308713), '063701': (33.6452131, -117.929843),
    '063702': (33.6470989, -117.9204613), '063802': (33.6758714, -117.9312877),
    '063803': (33.6702081, -117.9412986), '063806': (33.6799479, -117.9379014),
    '063807': (33.6637111, -117.9280607), '063808': (33.6550576, -117.9300385),
    '063902': (33.6781853, -117.9129197), '063903': (33.6848629, -117.9021898),
    '063904': (33.674841,  -117.8928079), '063905': (33.6655703, -117.9139178),
    '063909': (33.6902666, -117.8758095), '063910': (33.6938267, -117.9159916),
    '074006': (33.70141,   -117.8832581), '074102': (33.7193096, -117.8931177),
    '074103': (33.725167,  -117.8807777), '074107': (33.7001853, -117.9010592),
    '074108': (33.7210967, -117.9120597), '074109': (33.7170761, -117.8997703),
    '074111': (33.7103013, -117.8881087), '074405': (33.7482334, -117.8580242),
    '074406': (33.746832,  -117.8488154), '074407': (33.738184,  -117.8411724),
    '074408': (33.7385625, -117.8357717), '074501': (33.7356812, -117.855292),
    '074502': (33.7250125, -117.8597465), '074601': (33.7414062, -117.8686772),
    '074602': (33.7420004, -117.8648553), '074701': (33.7309928, -117.8912536),
    '074702': (33.7290569, -117.8807986), '074801': (33.7400922, -117.8893705),
    '074803': (33.7310683, -117.9044714), '074805': (33.737477,  -117.9012214),
    '074806': (33.7373513, -117.8961789), '074901': (33.7466803, -117.8807731),
    '075002': (33.7459929, -117.8704486), '075003': (33.754712,  -117.8744875),
    '075004': (33.7602313, -117.8655194), '075100': (33.7601445, -117.8783129),
    '075301': (33.7753362, -117.8770466), '075302': (33.7632085, -117.888511),
    '075303': (33.7655023, -117.8676904), '075403': (33.755891,  -117.8464954),
    '075404': (33.770435,  -117.8412211), '075405': (33.763558,  -117.8437837),
    '075504': (33.7508706, -117.8157432), '075505': (33.7415972, -117.8236884),
    '075506': (33.7395838, -117.803716),  '075507': (33.7334134, -117.8122285),
    '075512': (33.7327451, -117.8245087), '075513': (33.7270328, -117.8162436),
    '075514': (33.7363213, -117.8283098), '075516': (33.6936508, -117.8327198),
    '075517': (33.7288943, -117.8257066), '075604': (33.784612,  -117.7867056),
    '075605': (33.7848364, -117.7660795), '075606': (33.741525,  -117.7933162),
    '075607': (33.7341481, -117.7873916), '075701': (33.7637616, -117.8332816),
    '075702': (33.7564523, -117.8172982), '075703': (33.769969,  -117.8083904),
    '075805': (33.7913647, -117.8334487), '075806': (33.7790463, -117.8363681),
    '075807': (33.7835883, -117.8147047), '075808': (33.7768963, -117.8248077),
    '075809': (33.8206523, -117.8053061), '075810': (33.8138023, -117.8095508),
    '075811': (33.8180754, -117.8340706), '075812': (33.8051951, -117.8244345),
    '075813': (33.8272721, -117.8345885), '075815': (33.7922912, -117.8279431),
    '075901': (33.7896819, -117.8487286), '075902': (33.775668,  -117.8496817),
    '076001': (33.7824523, -117.8636747), '076002': (33.7756099, -117.8661691),
    '076102': (33.7835125, -117.8936753), '076104': (33.7948681, -117.8879017),
    '076105': (33.7910054, -117.8701864), '076201': (33.8326642, -117.8500395),
    '076202': (33.8330026, -117.8420821), '076204': (33.8092004, -117.8575436),
    '076205': (33.8057181, -117.8487062), '076206': (33.7970862, -117.8400036),
    '086303': (33.8105054, -117.8801786), '086304': (33.83445,   -117.8857145),
    '086305': (33.8293973, -117.872727),  '086306': (33.8190215, -117.884911),
    '086402': (33.8447361, -117.8857397), '086404': (33.8491216, -117.9021072),
    '086405': (33.8431049, -117.8922513), '086407': (33.844588,  -117.8694522),
    '086501': (33.8475582, -117.9168546), '086502': (33.8426667, -117.9082076),
    '086601': (33.8489095, -117.9395251), '086602': (33.8454665, -117.9369848),
    '086701': (33.8508709, -117.9566607), '086702': (33.8407726, -117.9444641),
    '086802': (33.8415924, -117.9634053), '086803': (33.8330097, -117.9825407),
    '086901': (33.8305449, -118.0087808), '086902': (33.8191875, -117.9982823),
    '086903': (33.8302599, -117.9917307), '087001': (33.8212277, -117.978071),
    '087002': (33.8212734, -117.9739221), '087101': (33.8361041, -117.963356),
    '087102': (33.8376403, -117.9498851), '087103': (33.8307933, -117.9542955),
    '087106': (33.8199072, -117.9305971), '087200': (33.8362762, -117.9357705),
    '087301': (33.8377877, -117.9159575), '087302': (33.8334383, -117.9178917),
    '087401': (33.8291345, -117.9107913), '087403': (33.8106381, -117.9081356),
    '087405': (33.8209872, -117.9187057), '087505': (33.8142946, -117.9305586),
    '087601': (33.8124162, -117.9370517), '087602': (33.7995653, -117.9370037),
    '087701': (33.8195584, -117.9611065), '087703': (33.8086292, -117.9632293),
    '087704': (33.8164911, -117.9567278), '087801': (33.804633,  -118.0063501),
    '087802': (33.8136998, -117.9971307), '087806': (33.8149547, -117.9712694),
    '087901': (33.7992822, -117.9962017), '088001': (33.8001904, -117.9792267),
    '088002': (33.7946904, -117.9704926), '088104': (33.7856762, -117.9947626),
    '088105': (33.7848135, -117.9797353), '088106': (33.7774672, -117.988401),
    '088302': (33.7850298, -117.9455001), '088401': (33.7924409, -117.9256799),
    '088402': (33.7852311, -117.9192724), '088403': (33.7852556, -117.9105137),
    '088501': (33.7779405, -117.9279205), '088502': (33.7780545, -117.9213869),
    '088601': (33.7777837, -117.943189),  '088602': (33.7706033, -117.9334553),
    '088701': (33.7794598, -117.9628871), '088702': (33.7704392, -117.9517548),
    '088801': (33.7699976, -117.9747063), '088802': (33.7630236, -117.9745881),
    '088901': (33.7540832, -117.9592443), '088902': (33.7619707, -117.9457672),
    '088903': (33.757761,  -117.9484488), '088904': (33.7474342, -117.9496277),
    '088905': (33.7540056, -117.9679933), '089003': (33.7619992, -117.9359371),
    '089004': (33.7506828, -117.9245602), '089102': (33.7634579, -117.9159425),
    '089104': (33.7570631, -117.9091556), '089105': (33.7471563, -117.9129908),
    '089106': (33.7720798, -117.9149277), '089107': (33.7711072, -117.8971375),
    '099202': (33.7346468, -117.9168328), '099203': (33.7349839, -117.9416819),
    '099204': (33.7422572, -117.9528509), '099212': (33.7140887, -117.9844808),
    '099215': (33.6777594, -117.958223),  '099220': (33.6530068, -117.9821137),
    '099222': (33.7430263, -117.9678523), '099224': (33.7195121, -117.9675349),
    '099225': (33.7197479, -117.95884),   '099226': (33.7272529, -117.9329544),
    '099227': (33.7198948, -117.9243763), '099229': (33.7017377, -117.936671),
    '099230': (33.6905362, -117.9497588), '099232': (33.6904648, -117.9649429),
    '099233': (33.7121184, -117.9761126), '099234': (33.7049649, -117.9695016),
    '099241': (33.7263053, -117.9841119), '099244': (33.6613555, -117.9753963),
    '099245': (33.6830996, -117.9843391), '099247': (33.7425735, -117.9354761),
    '099248': (33.7429814, -117.9267786), '099249': (33.7426065, -117.9136741),
    '099250': (33.7055649, -117.9593428), '099251': (33.7127806, -117.9518981),
    '099305': (33.6773464, -117.9953468), '099307': (33.6602154, -117.9908854),
    '099308': (33.6799606, -118.0229037), '099309': (33.6713746, -118.0009882),
    '099310': (33.6622086, -118.0116395), '099311': (33.658921,  -118.0007627),
    '099402': (33.7123102, -117.990879),  '099404': (33.7118075, -118.0195632),
    '099407': (33.7262747, -118.0284372), '099408': (33.718952,  -118.0371527),
    '099410': (33.7263277, -117.9961885), '099411': (33.7165861, -117.9934594),
    '099412': (33.702602,  -118.0108201), '099415': (33.7040658, -118.0292273),
    '099416': (33.7170504, -118.0479598), '099417': (33.7133711, -118.0457941),
    '099419': (33.6948522, -118.0026329), '099502': (33.7495409, -118.0652986),
    '099504': (33.7512108, -118.0946098), '099506': (33.7221327, -118.0822762),
    '099508': (33.7234587, -118.0462417), '099509': (33.7716326, -118.0795668),
    '099510': (33.7675069, -118.0914978), '099511': (33.7364488, -118.1065672),
    '099512': (33.7461891, -118.1077967), '099514': (33.7206557, -118.0603419),
    '099601': (33.7505402, -118.0084868), '099603': (33.737089,  -118.0372772),
    '099604': (33.7407881, -118.0200836), '099605': (33.7336782, -118.0068801),
    '099701': (33.7483016, -117.9917918), '099702': (33.7410983, -117.9851537),
    '099703': (33.7339353, -117.9764968), '099801': (33.7646353, -117.9940862),
    '099802': (33.7610056, -117.9939769), '099803': (33.7537547, -117.9794005),
    '099902': (33.7714189, -118.0209174), '099903': (33.7701686, -118.0117502),
    '099904': (33.7609386, -118.0052655), '110001': (33.7906599, -118.0360253),
    '110004': (33.7761971, -118.0282051), '110005': (33.7778797, -118.0382316),
    '110008': (33.7765971, -118.080983),  '110010': (33.7902676, -118.0149801),
    '110011': (33.7992957, -118.0154136), '110012': (33.7777481, -118.0622695),
    '110014': (33.8017856, -118.0654479), '110102': (33.8368208, -118.0405152),
    '110104': (33.8198875, -118.0412156), '110106': (33.8162542, -118.0589323),
    '110108': (33.8079105, -118.0765417), '110109': (33.8207018, -118.0313971),
    '110110': (33.8272763, -118.0332798), '110111': (33.8414019, -118.0550856),
    '110114': (33.8128513, -118.0500833), '110115': (33.8523519, -118.0441606),
    '110116': (33.8485836, -118.0329197), '110117': (33.8216563, -118.0571724),
    '110118': (33.8187063, -118.0493562), '110202': (33.8268699, -118.01299),
    '110203': (33.8134675, -118.0237754), '110301': (33.8547988, -118.0329324),
    '110302': (33.8496412, -118.0066607), '110303': (33.8428358, -118.0152287),
    '110304': (33.842868,  -118.0239143), '110401': (33.8520265, -117.9952067),
    '110402': (33.8357056, -117.9957087), '110500': (33.8655931, -118.0186975),
    '110603': (33.8681094, -117.9894607), '110604': (33.8910398, -117.9833775),
    '110605': (33.8810149, -117.9685179), '110607': (33.8797696, -117.9934015),
    # 11 estimated tracts
    '011714': (33.8820, -117.8972), '011722': (33.8698, -117.9198),
    '052411': (33.6742, -117.7762), '052432': (33.6441, -117.7311),
    '074403': (33.7523, -117.8831), '086301': (33.8431, -117.9801),
    '086801': (33.8391, -117.9251), '087105': (33.8231, -117.9021),
    '087504': (33.8031, -117.8701), '087902': (33.7641, -117.9751),
    '097702': (33.7231, -117.9751),
}

# ── SIDEBAR ────────────────────────────────────────────────────────────────────
st.sidebar.title("🏠 OC Housing Justice")
st.sidebar.markdown("*Data for Social Good*")
st.sidebar.markdown("---")

page = st.sidebar.radio("Navigate", [
    "📊 Overview", "🗺️ Interactive Map", "🚨 Underserved Tracts",
    "🏙️ City Analysis", "📈 Correlations", "🔍 Explore Data", "📋 Methodology"
])

st.sidebar.markdown("---")
st.sidebar.markdown(f"**Programs:** {len(df):,}")
st.sidebar.markdown(f"**Total Beds:** {int(df['Total Beds'].sum()):,}")
st.sidebar.markdown(f"**Cities:** {df['City'].nunique()}")
st.sidebar.markdown(f"**Tracts Analyzed:** {tract_agg.shape[0]}")
st.sidebar.markdown(f"**Underserved Tracts:** {int(tract_agg['Underserved'].sum())}")
st.sidebar.markdown("---")
st.sidebar.markdown("**By:** Paxton Wang")
st.sidebar.markdown("**Data:** HIC HDX 2025 · CalEnviroScreen 4.0 · Census ACS")

# ── OVERVIEW ───────────────────────────────────────────────────────────────────
if page == "📊 Overview":
    st.title("🏠 Orange County Housing Justice Dashboard")
    st.markdown("### Mapping the gap between housing need and resource availability")
    st.markdown("---")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-value">{len(df):,}</div>
            <div class="metric-label">Total Housing Programs</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-value">{int(df['Total Beds'].sum()):,}</div>
            <div class="metric-label">Total Beds Across OC</div>
        </div>""", unsafe_allow_html=True)
    with c3:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-value" style="color:#D95A00">{int(tract_agg['Underserved'].sum())}</div>
            <div class="metric-label">Critically Underserved Tracts</div>
        </div>""", unsafe_allow_html=True)
    with c4:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-value">{df['City'].nunique()}</div>
            <div class="metric-label">Cities Analyzed</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        st.subheader("Need Score Distribution")
        fig = px.histogram(tract_agg, x='Need_Score', nbins=15,
                           color_discrete_sequence=['#1A5276'],
                           labels={'Need_Score': 'Composite Need Score'})
        fig.add_vline(x=tract_agg['Need_Score'].median(), line_dash="dash",
                      line_color="#D95A00",
                      annotation_text=f"Median: {tract_agg['Need_Score'].median():.3f}",
                      annotation_font_color="#1A2B3C")
        st.plotly_chart(style_fig(fig), use_container_width=True)

    with c2:
        st.subheader("Beds vs. Need Score by Tract")
        fig = px.scatter(tract_agg, x='Need_Score', y='Total_Beds',
                         color='Underserved',
                         color_discrete_map={True: '#D95A00', False: '#1A5276'},
                         hover_data=['Census Tract', 'City'],
                         labels={'Need_Score': 'Need Score', 'Total_Beds': 'Total Beds'},
                         size='Total_Beds', size_max=40)
        fig.add_hline(y=tract_agg['Total_Beds'].median(), line_dash="dash",
                      line_color="gray", opacity=0.5, annotation_text="Median beds",
                      annotation_font_color="#1A2B3C")
        fig.add_vline(x=tract_agg['Need_Score'].median(), line_dash="dash",
                      line_color="gray", opacity=0.5, annotation_text="Median need",
                      annotation_font_color="#1A2B3C")
        st.plotly_chart(style_fig(fig), use_container_width=True)

    st.markdown("""<div class="insight-box">
        <strong>Key Finding:</strong> Areas with higher housing need scores tend to have fewer beds —
        the correlation between rent burden and total beds is −0.243, confirming resources
        are misaligned with community need.
    </div>""", unsafe_allow_html=True)

    st.subheader("Program Types Across OC")
    proj_counts = df['Project Type'].value_counts().reset_index()
    proj_counts.columns = ['Project Type', 'Count']
    fig = px.bar(proj_counts, x='Project Type', y='Count',
                 color='Count', color_continuous_scale='Blues')
    fig.update_layout(showlegend=False)
    st.plotly_chart(style_fig(fig), use_container_width=True)

# ── MAP ────────────────────────────────────────────────────────────────────────
elif page == "🗺️ Interactive Map":
    st.title("🗺️ Interactive Housing Need Map")
    st.markdown("Each circle = one census tract. **Size** = total beds. **Color** = composite need score.")
    st.markdown("---")

    c1, c2 = st.columns(2)
    with c1:
        min_need = st.slider("Minimum Need Score", 0.0, 1.0, 0.0, 0.05)
    with c2:
        show_under_only = st.checkbox("Show underserved tracts only", False)

    st.markdown("**Map Layers:**")
    lc1, lc2, lc3, lc4 = st.columns(4)
    with lc1: layer_enviro   = st.checkbox("🌿 CalEnviroScreen", True)
    with lc2: layer_rent     = st.checkbox("🏠 Rent Burden", False)
    with lc3: layer_poverty  = st.checkbox("📉 Poverty Rate", False)
    with lc4: layer_programs = st.checkbox("📍 Program Pins", True)

    st.markdown("---")
    l1, l2, l3, l4, l5 = st.columns(5)
    with l1: st.markdown("🟠 **Critical** (≥0.80)")
    with l2: st.markdown("🟡 **High** (≥0.70)")
    with l3: st.markdown("🔵 **Moderate** (≥0.60)")
    with l4: st.markdown("🔷 **Lower** (<0.60)")
    with l5: st.markdown("⚠️ **Underserved**")

    filtered = tract_agg[tract_agg['Need_Score'] >= min_need]
    if show_under_only:
        filtered = filtered[filtered['Underserved']]

    m = folium.Map(location=[33.74, -117.87], zoom_start=10, tiles='CartoDB positron')
    lg_need        = folium.FeatureGroup(name="Need Score", show=True)
    lg_enviro      = folium.FeatureGroup(name="🌿 CalEnviroScreen", show=layer_enviro)
    lg_rent        = folium.FeatureGroup(name="🏠 Rent Burden", show=layer_rent)
    lg_poverty     = folium.FeatureGroup(name="📉 Poverty Rate", show=layer_poverty)
    lg_programs    = folium.FeatureGroup(name="📍 Program Pins", show=layer_programs)
    lg_underserved = folium.FeatureGroup(name="⚠️ Underserved", show=True)

    mapped = 0
    for _, row in filtered.iterrows():
        tract = clean_tract(row['Census Tract'])
        if tract not in TRACT_COORDS:
            continue
        lat, lon = TRACT_COORDS[tract]
        col = need_color(row['Need_Score'])
        rad = max(8, min(40, row['Total_Beds'] / 10))
        mapped += 1

        ev = row.get('Enviro_Percentile')
        ev_str = f"{ev:.0%}" if ev and ev == ev else "N/A"

        popup_html = f"""
        <div style="font-family:Arial;width:240px;line-height:1.5;color:#1A2B3C">
            <h4 style="color:#1A5276;margin:0 0 6px 0;border-bottom:2px solid #1A5276;padding-bottom:4px">
                Tract {tract} — {row['City']}</h4>
            <table style="width:100%;font-size:12px">
                <tr><td><b>Need Score</b></td>
                    <td style="color:{col};font-weight:bold;font-size:14px">{row['Need_Score']:.3f}</td></tr>
                <tr><td><b>Total Beds</b></td><td>{int(row['Total_Beds'])}</td></tr>
                <tr><td><b># Programs</b></td><td>{int(row['Num_Programs'])}</td></tr>
                <tr><td><b>Rent Burden</b></td><td>{row['Rent_Burden']:.0%}</td></tr>
                <tr><td><b>Poverty %ile</b></td><td>{row['Poverty_Percentile']:.0%}</td></tr>
                <tr><td><b>Income %ile</b></td><td>{row['Income_Percentile']:.0%}</td></tr>
                <tr><td><b>CalEnviroScreen</b></td><td>{ev_str}</td></tr>
            </table>
            {"<div style='margin-top:8px;padding:4px 8px;background:#FDDBB8;border-radius:4px;color:#D95A00;font-weight:bold'>⚠️ UNDERSERVED TRACT</div>" if row['Underserved'] else ""}
        </div>"""

        folium.CircleMarker(
            location=[lat, lon], radius=rad,
            color=col, fill=True, fill_color=col, fill_opacity=0.75, weight=2,
            popup=folium.Popup(popup_html, max_width=260),
            tooltip=f"Tract {tract} | Need: {row['Need_Score']:.3f} | Beds: {int(row['Total_Beds'])}"
        ).add_to(lg_need)

        if layer_rent:
            rb = row.get('Rent_Burden')
            if rb and rb == rb:
                folium.CircleMarker(
                    location=[lat - 0.003, lon + 0.003], radius=max(5, rad * 0.6),
                    color=lerp_color(min(rb * 1.5, 1.0), (52,152,219), (142,68,173)),
                    fill=True, fill_opacity=0.5, weight=1, dash_array='6',
                    tooltip=f"Rent Burden: {rb:.0%}"
                ).add_to(lg_rent)

        if layer_poverty:
            pov = row.get('Poverty_Percentile')
            if pov and pov == pov:
                folium.CircleMarker(
                    location=[lat - 0.003, lon - 0.003], radius=max(5, rad * 0.6),
                    color=lerp_color(pov, (243,156,18), (192,57,43)),
                    fill=True, fill_opacity=0.5, weight=1, dash_array='8',
                    tooltip=f"Poverty %ile: {pov:.0%}"
                ).add_to(lg_poverty)

        if row['Underserved']:
            folium.Marker(
                location=[lat + 0.006, lon],
                icon=folium.DivIcon(
                    html='<div style="font-size:18px;text-shadow:1px 1px 2px white">⚠️</div>',
                    icon_size=(20, 20), icon_anchor=(10, 10)
                ),
                tooltip=f"⚠️ Underserved: {tract}"
            ).add_to(lg_underserved)

    if layer_programs:
        prog_df = enriched.copy()
        prog_df['Census Tract'] = prog_df['Census Tract'].apply(clean_tract)
        for _, prog in prog_df.iterrows():
            t = prog.get('Census Tract')
            if t and t in TRACT_COORDS:
                plat, plon = TRACT_COORDS[t]
                random.seed(hash(str(prog.get('Project Name', ''))))
                jlat = plat + random.uniform(-0.008, 0.008)
                jlon = plon + random.uniform(-0.008, 0.008)
                beds  = prog.get('Total Beds', 0)
                ptype = prog.get('Project Type', '')
                pin_col = {
                    'Emergency Shelter': 'red',
                    'Transitional Housing': 'orange',
                    'Permanent Supportive Housing ': 'blue',
                    'Other Permanent Housing': 'darkblue',
                    'Rapid Re-Housing': 'green',
                }.get(ptype, 'gray')
                folium.CircleMarker(
                    location=[jlat, jlon], radius=4,
                    color='white', fill=True, fill_color=pin_col, fill_opacity=0.9, weight=1.5,
                    tooltip=f"{prog.get('Organization Name','')} | {ptype} | {int(beds) if beds==beds else '?'} beds",
                    popup=folium.Popup(
                        f"<b>{prog.get('Project Name','')}</b><br>"
                        f"{prog.get('Organization Name','')}<br>"
                        f"Type: {ptype}<br>City: {prog.get('City','')}<br>"
                        f"Beds: {int(beds) if beds==beds else 'Unknown'}",
                        max_width=200)
                ).add_to(lg_programs)

    # CalEnviroScreen 5.0 polygon choropleth
    if layer_enviro and tract_geojson is not None:

        def es_style(feature):
            # Try GEOID field first (11-digit), then build from TRACTCE (6-digit)
            props = feature['properties']
            geoid = props.get('GEOID', '')
            if not geoid:
                state  = props.get('STATEFP', '06')
                county = props.get('COUNTYFP', '059')
                tract  = props.get('TRACTCE', '')
                geoid  = f"{state}{county}{tract}"
            val = ces50.get(geoid)
            if val is None:
                return {'fillColor': '#DDDDDD', 'color': '#BBBBBB', 'weight': 0.4, 'fillOpacity': 0.15}
            return {
                'fillColor': enviro_color(val),
                'color': '#555555',
                'weight': 0.8,
                'fillOpacity': 0.40,
            }

        def es_highlight(feature):
            return {'weight': 2.5, 'color': '#1A5276', 'fillOpacity': 0.80}

        def es_tooltip(feature):
            props = feature['properties']
            geoid = props.get('GEOID', '')
            if not geoid:
                geoid = f"06059{props.get('TRACTCE', '')}"
            val   = ces50.get(geoid)
            pct_str = f"{val*100:.1f}th percentile" if val is not None else "No data"
            name  = props.get('NAMELSAD', props.get('NAME', geoid))
            return f"{name} — CES 5.0: {pct_str}"

        folium.GeoJson(
            tract_geojson,
            style_function=es_style,
            highlight_function=es_highlight,
            tooltip=folium.GeoJsonTooltip(
                fields=['GEOID'] if 'GEOID' in (tract_geojson['features'][0]['properties'] if tract_geojson['features'] else {}) else ['TRACTCE'],
                aliases=['Tract GEOID'],
                localize=True,
            ),
        ).add_to(lg_enviro)

        if ces50:
            st.caption(f"🌿 CalEnviroScreen 5.0 — {len(ces50)} OC tracts colored by cumulative burden percentile")
        else:
            st.warning(f"⚠️ ces50_oc.json not found. Script: `{os.path.abspath(__file__)}` | cwd: `{os.getcwd()}`")

    elif layer_enviro and tract_geojson is None:
        st.warning("⚠️ Place `oc_tracts.geojson` in the same folder as this script to enable polygon boundaries.")

    # Layer order matters — add enviro polygons FIRST so circles render on top
    for lg in [lg_enviro, lg_rent, lg_poverty, lg_need, lg_programs, lg_underserved]:
        lg.add_to(m)
    folium.LayerControl(collapsed=False).add_to(m)

    st_folium(m, width=None, height=580)
    st.markdown(f"**{mapped} of {len(filtered)} tracts mapped**")
    if layer_programs:
        st.markdown("**Program pins:** 🔴 Emergency Shelter · 🟠 Transitional · 🔵 Permanent Supportive · 🟢 Rapid Re-Housing")

# ── UNDERSERVED ────────────────────────────────────────────────────────────────
elif page == "🚨 Underserved Tracts":
    st.title("🚨 Critically Underserved Tracts")
    st.markdown("Above-median need + below-median beds = most urgent gaps in OC's housing safety net.")
    st.markdown("---")

    underserved = tract_agg[tract_agg['Underserved']].sort_values('Need_Score', ascending=False)
    st.markdown(f"**{len(underserved)} underserved tracts** out of {len(tract_agg)} analyzed")

    for _, row in underserved.iterrows():
        col    = need_color(row['Need_Score'])
        ev     = row['Enviro_Percentile']
        ev_str = f"{ev:.0%}" if ev == ev else "N/A"
        st.markdown(f"""
        <div class="alert-box">
            <h4 style="margin:0;color:#D95A00">⚠️ Tract {row['Census Tract']} — {row['City']}</h4>
            <div style="display:flex;gap:24px;margin-top:10px;flex-wrap:wrap">
                <div><b>Need Score:</b> <span style="color:{col};font-weight:700;font-size:1.2rem">{row['Need_Score']:.3f}</span></div>
                <div><b>Beds:</b> {int(row['Total_Beds'])}</div>
                <div><b>Rent Burden:</b> {row['Rent_Burden']:.0%}</div>
                <div><b>Poverty %ile:</b> {row['Poverty_Percentile']:.0%}</div>
                <div><b>Income %ile:</b> {row['Income_Percentile']:.0%}</div>
                <div><b>Enviro %ile:</b> {ev_str}</div>
            </div>
        </div>""", unsafe_allow_html=True)

    st.markdown("---")
    fig = px.bar(underserved, x='Census Tract', y=['Need_Score', 'Total_Beds'],
                 barmode='group', color_discrete_sequence=['#D95A00', '#1A5276'])
    st.plotly_chart(style_fig(fig), use_container_width=True)

# ── CITY ANALYSIS ──────────────────────────────────────────────────────────────
elif page == "🏙️ City Analysis":
    st.title("🏙️ City-Level Analysis")
    st.markdown("---")

    all_cities = sorted(city_summary['City'].unique())
    sel_cities = st.multiselect("Filter cities", all_cities, default=all_cities)
    city_f = city_summary[city_summary['City'].isin(sel_cities)]

    c1, c2 = st.columns(2)
    with c1:
        st.subheader("Avg Need Score by City")
        cs  = city_f.sort_values('Avg_Need_Score', ascending=True)
        fig = go.Figure(go.Bar(x=cs['Avg_Need_Score'], y=cs['City'],
                               orientation='h',
                               marker_color=[need_color(s) for s in cs['Avg_Need_Score']]))
        fig.update_layout(xaxis_title='Avg Need Score', height=max(300, len(cs) * 28))
        st.plotly_chart(style_fig(fig), use_container_width=True)

    with c2:
        st.subheader("Total Beds by City")
        cb  = city_f.sort_values('Total_Beds', ascending=True)
        fig = go.Figure(go.Bar(x=cb['Total_Beds'], y=cb['City'],
                               orientation='h', marker_color='#1A5276'))
        fig.update_layout(xaxis_title='Total Beds', height=max(300, len(cb) * 28))
        st.plotly_chart(style_fig(fig), use_container_width=True)

    st.subheader("Need vs. Beds — City Scatter")
    fig = px.scatter(city_f, x='Avg_Need_Score', y='Total_Beds',
                     text='City', size='Num_Programs', color='Avg_Need_Score',
                     color_continuous_scale=['#2E86C1', '#F0A500', '#E8870A', '#D95A00'])
    fig.update_traces(textposition='top center')
    fig.update_layout(height=500)
    st.plotly_chart(style_fig(fig), use_container_width=True)

    st.subheader("Full City Table")
    disp = city_f.sort_values('Avg_Need_Score', ascending=False).copy()
    disp.columns = ['City', 'Total Beds', 'Avg Need Score', '# Programs', 'Avg Rent Burden', 'Avg Poverty %ile']
    st.dataframe(disp, use_container_width=True, hide_index=True)

# ── CORRELATIONS ───────────────────────────────────────────────────────────────
elif page == "📈 Correlations":
    st.title("📈 Correlation Analysis")
    st.markdown("---")

    corr_cols   = ['Rent Burden %', 'Income Need', 'Poverty Line Percentile', 'Total Beds']
    corr_labels = ['Rent Burden', 'Income Need', 'Poverty %ile', 'Total Beds']
    corr_matrix = enriched[corr_cols].corr().round(3)

    c1, c2 = st.columns(2)
    with c1:
        st.subheader("Correlation Heatmap")
        fig = px.imshow(corr_matrix, x=corr_labels, y=corr_labels,
                        color_continuous_scale='RdBu_r', zmin=-1, zmax=1, text_auto=True)
        fig.update_layout(height=420)
        fig.update_traces(textfont=dict(color='#1A2B3C', size=13))
        fig.update_coloraxes(colorbar=dict(
            tickfont=dict(color='#1A2B3C'),
            title_font=dict(color='#1A2B3C'),
        ))
        st.plotly_chart(style_fig(fig), use_container_width=True)

    with c2:
        st.subheader("Rent Burden vs. Poverty Rate")
        fig = px.scatter(enriched, x='Rent Burden %', y='Poverty Line Percentile',
                         color='Need Score',
                         color_continuous_scale=['#2E86C1', '#AED6F1', '#FDDBB8', '#D95A00'],
                         hover_data=['City', 'Census Tract'], trendline='ols')
        fig.update_layout(height=420)
        st.plotly_chart(style_fig(fig), use_container_width=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""<div class="insight-box"><b>Income &amp; Poverty (r=0.677):</b>
          Strong — both measure economic deprivation from different angles.</div>""",
                    unsafe_allow_html=True)
        st.markdown("""<div class="insight-box"><b>Rent Burden &amp; Income (r=0.344):</b>
          Moderate — rent stress is partly but not fully income-driven.</div>""",
                    unsafe_allow_html=True)
    with c2:
        st.markdown("""<div class="insight-box"><b>Rent Burden &amp; Beds (r=−0.243):</b>
          Negative — highest need areas have fewest beds.</div>""",
                    unsafe_allow_html=True)
        st.markdown("""<div class="insight-box"><b>CalEnviroScreen — map layer only:</b>
          Removed from formula to avoid double-counting population vulnerability.</div>""",
                    unsafe_allow_html=True)

# ── EXPLORE DATA ───────────────────────────────────────────────────────────────
elif page == "🔍 Explore Data":
    st.title("🔍 Explore the Full Dataset")
    st.markdown("---")

    c1, c2, c3 = st.columns(3)
    with c1:
        sel_city = st.selectbox("City", ["All"] + sorted(df['City'].dropna().unique().tolist()))
    with c2:
        sel_type = st.selectbox("Project Type", ["All"] + sorted(df['Project Type'].dropna().unique().tolist()))
    with c3:
        min_beds_val = st.number_input("Min Beds", min_value=0, value=0, step=5)

    fdf = df.copy()
    if sel_city != "All":
        fdf = fdf[fdf['City'] == sel_city]
    if sel_type != "All":
        fdf = fdf[fdf['Project Type'] == sel_type]
    fdf = fdf[pd.to_numeric(fdf['Total Beds'], errors='coerce').fillna(0) >= min_beds_val]

    st.markdown(f"**{len(fdf)} programs** match your filters")
    disp_cols = ['Organization Name', 'Project Name', 'City', 'Project Type',
                 'Total Beds', 'Census Tract', 'Rent Burden %', 'Poverty Line Percentile']
    st.dataframe(fdf[[c for c in disp_cols if c in fdf.columns]],
                 use_container_width=True, hide_index=True)
    st.download_button(
        label="⬇️ Download filtered data as CSV",
        data=fdf.to_csv(index=False).encode('utf-8'),
        file_name='oc_housing_filtered.csv',
        mime='text/csv'
    )

# ── METHODOLOGY ────────────────────────────────────────────────────────────────
elif page == "📋 Methodology":
    st.title("📋 Methodology")
    st.markdown("---")

    st.subheader("Data Sources")
    st.markdown(f"""
| Source | Variables | Description |
|--------|-----------|-------------|
| HIC HDX 2025 (HUD) | Program type, beds, city, tract | {len(df)} OC housing programs |
| CalEnviroScreen 4.0 | Environmental burden %ile | CA EPA cumulative burden scores |
| Census ACS | Median HH income %ile | Tract-level income data |
| HUD CHAS | Rent burden %, poverty %ile | Housing affordability data |
    """)

    st.subheader("Composite Need Score Formula")
    st.code("""Need Score = (Rent Burden × 0.40) + ((1 − Income Percentile) × 0.25) + (Poverty Percentile × 0.35)

CalEnviroScreen 4.0 is shown as a separate map toggle layer only.
Excluded from formula to avoid double-counting population vulnerability.""")

    med_need = tract_agg['Need_Score'].median()
    med_beds = tract_agg['Total_Beds'].median()
    st.subheader("Underserved Classification")
    st.markdown(f"""
A tract is **underserved** when:
- Need Score > **{med_need:.3f}** (dataset median)
- Total Beds < **{med_beds:.0f}** (dataset median)
    """)

    enriched_pct = len(enriched) / len(df) * 100
    st.subheader("Limitations")
    st.markdown(f"""
- Only {len(enriched)} of {len(df)} programs ({enriched_pct:.0f}%) have census tract data
- Static 2025 snapshot — housing need changes over time
- Bed counts do not reflect vacancy rates or waitlists
- Program coordinates are census tract centroids (approximate), not exact addresses
- Tract boundaries: 471 from Census TIGERweb 2020; 11 estimated
    """)

    st.subheader("About")
    st.markdown("""
Built by **Paxton Wang** as part of the Orange County Housing Justice Initiative —
using data science to identify gaps in housing resources and inform county-level policy.

**Contact:** paxtonmwang@gmail.com
    """)
