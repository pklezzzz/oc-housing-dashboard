"""
Adds a CalEnviroScreen 5.0 percentile column to the 'Short' sheet of
HousingCentersOCFinalData.xlsx, using the tract -> percentile lookup in
ces50_oc.json.

NOTE ON THE OLD VERSION OF THIS SCRIPT: format_oc_tract() built GEOIDs as
'6059' + 6-digit tract (10 digits total). Census GEOIDs are STATE(2) +
COUNTY(3) + TRACT(6) = 11 digits, i.e. '06059' + tract. Dropping the leading
zero off the state code meant almost every tract failed to match the
CalEnviroScreen lookup. Fixed below.
"""

import json
import pandas as pd
import openpyxl

XLSX_PATH = 'HousingCentersOCFinalData.xlsx'
SHEET_NAME = 'Short'
CES50_LOOKUP_PATH = 'ces50_oc.json'
OUTPUT_CSV = 'HousingCentersOC_CES50_Updated.csv'


def format_oc_tract_to_geoid(tract_val):
    """Convert a 6-digit OC tract code (e.g. '087505') into an 11-digit
    Census GEOID (e.g. '06059087505')."""
    if pd.isna(tract_val):
        return None
    val_str = str(tract_val).split('.')[0].strip().zfill(6)
    return f"06059{val_str}"


# 1. Load the CES 5.0 tract -> percentile lookup
with open(CES50_LOOKUP_PATH, 'r') as f:
    ces_lookup = json.load(f)

# 2. Load the 'Short' sheet (headers are on row 1 in this workbook)
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb[SHEET_NAME]
headers = [c.value for c in ws[1]]
data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
housing_df = pd.DataFrame(data, columns=headers)
housing_df = housing_df.loc[:, housing_df.columns.notna()]
housing_df = housing_df.loc[:, ~housing_df.columns.duplicated()]

# 3. Build the GEOID and map in CES 5.0 percentiles as a NEW column
#    (left alongside the existing 'CalEnviroScreen 4.0 Percentile' column
#    rather than overwriting it, so both scores stay auditable)
housing_df['GEOID'] = housing_df['Census Tract'].apply(format_oc_tract_to_geoid)
housing_df['CalEnviroScreen 5.0 Percentile'] = housing_df['GEOID'].map(ces_lookup)
housing_df.drop(columns=['GEOID'], inplace=True)

matched = housing_df['CalEnviroScreen 5.0 Percentile'].notna().sum()
print(f"Matched {matched} of {len(housing_df)} properties to a CES 5.0 tract percentile")

# 4. Save to CSV
housing_df.to_csv(OUTPUT_CSV, index=False)
print(f"Updated file successfully created: {OUTPUT_CSV}")
